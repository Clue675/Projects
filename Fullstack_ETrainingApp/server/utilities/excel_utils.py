from openpyxl import load_workbook, Workbook
import os

def read_from_excel(query):
    wb = load_workbook(os.getenv('EXCEL_FILE_PATH'))
    ws = wb['EmployeeSheet1']
    results = []
    for row in ws.iter_rows(values_only=True):
        if query.lower() in str(row).lower():
            results.append({
                "First Name": row[0],
                "Last Name": row[1],
                "Badge Number": row[2],
                "Training Title": row[3],
                "Trainer": row[4],
                "Date": row[5],
                "Work Center": row[6]  # Placeholder or actual value
            })
    return results

def update_excel_with_new_training(training):
    wb = load_workbook(os.getenv('EXCEL_FILE_PATH'))
    ws = wb['EmployeeSheet1']
    new_row = [
        training['firstName'],
        training['lastName'],
        training['badgeNumber'],
        training.get('title', ""),  # Training Title
        training.get('trainer', ""),  # Trainer
        training['date'],
        ""  # Empty cell for Work Center
    ]
    ws.append(new_row)
    wb.save(os.getenv('EXCEL_FILE_PATH'))
