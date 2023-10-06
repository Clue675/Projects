from flask import Flask, jsonify, request, make_response, current_app
from flask_cors import CORS  # Import CORS
from openpyxl import load_workbook
import os
import logging

app = Flask(__name__)
CORS(app)  # Add CORS to the app
logging.basicConfig(level=logging.DEBUG)

# Use environment variable for Excel file path
excel_file_path = os.environ.get("EXCEL_FILE_PATH")

def add_training_data_to_excel(title, date, names):
    workbook = load_workbook(filename=excel_file_path)
    worksheet = workbook['EmployeeSheet1']
    next_row = worksheet.max_row + 1
    
    for name in names:
        worksheet.cell(row=next_row, column=1, value=name['first_name'])
        worksheet.cell(row=next_row, column=2, value=name['last_name'])
        worksheet.cell(row=next_row, column=3, value=name['badge_number'])
        worksheet.cell(row=next_row, column=4, value=title)
        worksheet.cell(row=next_row, column=5, value=date)
        next_row += 1
    
    workbook.save(filename=excel_file_path)

def read_from_excel(query):
    wb = load_workbook(excel_file_path)
    ws = wb['EmployeeSheet1']
    
    results = []
    for row in ws.iter_rows(values_only=True):
        if query.lower() in str(row).lower():
            results.append({
                "First Name": row[0],
                "Last Name": row[1],
                "Badge Number": row[2],
                "Work Center": row[3],
                "Training Title": row[4],
                "Trainer": row[5],
                "Date": row[6]
            })
    return results

@app.route('/addNewTraining', methods=['POST'])
def add_training():
    data = request.json
    title = data.get('title')
    date = data.get('date')
    names = data.get('names')
    
    add_training_data_to_excel(title, date, names)
    
    return jsonify({"message": "Training data added successfully"})

@app.route('/search', methods=['GET'])
def search_training():
    query = request.args.get('query')
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 50))
    start_index = (page - 1) * limit
    end_index = page * limit

    all_results = read_from_excel(query)
    total = len(all_results)

    results = all_results[start_index:end_index]

    return jsonify({
        "message": f"Search results for query: {query}",
        "data": results,
        "total": total,
        "page": page,
        "limit": limit
    })


@app.route('/previewData', methods=['POST'])
@jwt_required()
def preview_data():
    # Receive scanned data from the client
    data = request.json
    scanned_data = data.get('scannedData')

    # Process the scanned data
    # (assuming scanned_data is a list of dictionaries)
    processed_data = []
    for record in scanned_data:
        processed_data.append({
            "First Name": record.get("first_name"),
            "Last Name": record.get("last_name"),
            "Badge Number": record.get("badge_number"),
            "Training Title": record.get("title"),
            "Trainer": record.get("trainer"),
            "Date": record.get("date")
        })

    # Send the processed data back to the client
    return make_response(jsonify({"message": "Data processed successfully", "processedData": processed_data}), 200)


if __name__ == '__main__':
    app.run(debug=True, port=5000)
