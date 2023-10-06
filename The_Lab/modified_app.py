from flask import Flask, jsonify, request
from flask_cors import CORS
from flask_jwt_extended import JWTManager, jwt_required, create_access_token, get_jwt_identity
from passlib.hash import pbkdf2_sha256
from pymongo import MongoClient
from openpyxl import load_workbook
from docx2pdf import convert
from docx import Document
from dotenv import load_dotenv
from werkzeug.utils import secure_filename
import os

# Load environment variables
load_dotenv()

# Initialize Flask app and configurations
app = Flask(__name__)
CORS(app)

app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')
app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY')
app.config['UPLOAD_FOLDER'] = './Training Logs'
app.config['ALLOWED_EXTENSIONS'] = {'pdf', 'docx'}

# Initialize JWT
jwt = JWTManager(app)

# MongoDB Setup
client = MongoClient('mongodb://localhost:27017/')
db = client['my_database']
users = db['users']
# Checks if uploaded file has allowed extension
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# Reads data from Excel based on a query
def read_from_excel(query):
    wb = load_workbook(os.getenv('EXCEL_FILE_PATH'))
    ws = wb['EmployeeSheet1']
    results = []
    for row in ws.iter_rows(values_only=True):
        if query.lower() in str(row).lower():
            results.append({
                "First Name": row[0],
                "Last Name": row[1],
                "Badge Number": row[2],
                "Work Center": row[3],
                "Training Title": row[4],
                "Trainer": row[5],
                "Date": row[6]
            })
    return results

# Adds training data to an Excel sheet
def add_training_data_to_excel(title, date, trainer, names):
    try:
        excel_file_path = os.getenv('EXCEL_FILE_PATH')
        workbook = load_workbook(filename=excel_file_path)
        worksheet = workbook['EmployeeSheet1']
        next_row = worksheet.max_row + 1
        for name in names:
            worksheet.cell(row=next_row, column=1, value=name['first_name'])
            worksheet.cell(row=next_row, column=2, value=name['last_name'])
            worksheet.cell(row=next_row, column=3, value=name['badge_number'])
            worksheet.cell(row=next_row, column=4, value=title)
            worksheet.cell(row=next_row, column=5, value=trainer)
            worksheet.cell(row=next_row, column=6, value=date)
            next_row += 1
        workbook.save(filename=excel_file_path)
        return True
    except Exception as e:
        print(f"An error occurred: {e}")
        return False
# Registration endpoint
@app.route('/register', methods=['POST'])
def register():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    hashed_password = pbkdf2_sha256.hash(password)
    users.insert_one({"username": username, "password": hashed_password})
    return jsonify({"message": "User registered successfully"})

# Login endpoint
@app.route('/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    user = users.find_one({"username": username})
    if pbkdf2_sha256.verify(password, user['password']):
        access_token = create_access_token(identity=username)
        return jsonify({"message": "Login successful", "access_token": access_token})
    else:
        return jsonify({"message": "Invalid credentials"}), 401
# File Upload endpoint
@app.route('/upload', methods=['POST'])
@jwt_required()
def upload_file():
    current_user = get_jwt_identity()
    title = request.form.get("title")
    date = request.form.get("date")
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)
    extracted_data = []
    # Add your own logic to parse the extracted text and populate extracted_data
    return jsonify({"message": "File uploaded and processed", "previewData": extracted_data})

# Add New Training endpoint
@app.route('/addNewTraining', methods=['POST'])
@jwt_required()
def add_new_training():
    data = request.json
    title = data.get('title')
    date = data.get('date')
    trainer = data.get('trainer')
    names = data.get('names')
    add_training_data_to_excel(title, date, trainer, names)
    return jsonify({"message": "Training data added successfully"})

# Search endpoint
@app.route('/search', methods=['GET'])
@jwt_required()
def search_training():
    query = request.args.get('query')
    results = read_from_excel(query)
    return jsonify({"message": f"Search results for query: {query}", "data": results})

# Main execution
if __name__ == '__main__':
    app.run(debug=True, port=5000)
