from datetime import timedelta
from collections import defaultdict
from flask import Flask, request, session, jsonify, redirect, url_for, make_response, current_app
from flask_cors import CORS, cross_origin
import openpyxl
from datetime import datetime  # Importing datetime
from flask_jwt_extended import JWTManager, jwt_required, create_access_token, get_jwt_identity
from flask_sqlalchemy import SQLAlchemy
from passlib.hash import pbkdf2_sha256
from pymongo import MongoClient
from dotenv import load_dotenv
from flask_migrate import Migrate
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
from collections import defaultdict
from bson import ObjectId
import pandas as pd
import pythoncom
import os
import logging
import re
import traceback


# Load environment variables
load_dotenv()

# Initialize Flask app
app = Flask(__name__)
CORS(app)

# Configuration
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')
app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY')
app.config['UPLOAD_FOLDER'] = './Training Logs'
app.config['ALLOWED_EXTENSIONS'] = {'pdf', 'docx'}
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///trainings.db'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=5)  # 5-hour session

# fallback to 'Employee_database.xlsx'
EXCEL_FILE_PATH = os.getenv('EXCEL_FILE_PATH', 'Employee_database.xlsx')

# Initialize SQLAlchemy
sql_db = SQLAlchemy(app)

# Initialize Migrate
migrate = Migrate(app, sql_db)

# Initialize JWT
jwt = JWTManager(app)
jwt._set_error_handler_callbacks(app)


@app.before_request
def log_request_info():
    app.logger.debug('Headers: %s', request.headers)
    app.logger.debug('Body: %s', request.get_data())


# MongoDB setup
client = MongoClient('mongodb://localhost:27017/')
mongo_db = client['my_database']
users = mongo_db['users']
users_collection = users


# SQLAlchemy model for Training
class Training(sql_db.Model):
    id = sql_db.Column(sql_db.Integer, primary_key=True)
    firstName = sql_db.Column(sql_db.String(50))
    lastName = sql_db.Column(sql_db.String(50))
    badgeNumber = sql_db.Column(sql_db.String(50))
    title = sql_db.Column(sql_db.String(100))
    date = sql_db.Column(sql_db.String(50))
    trainer = sql_db.Column(sql_db.String(50))

    @property
    def serialize(self):
        return {
            'id': self.id,
            'firstName': self.firstName,
            'lastName': self.lastName,
            'badgeNumber': self.badgeNumber,
            'title': self.title,
            'date': self.date,
            'trainer': self.trainer
        }


def read_from_excel(query):
    wb = load_workbook(EXCEL_FILE_PATH)
    ws = wb.active
    query_fields = ["firstName", "lastName", "badgeNumber", "title", "trainer", "date"]
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all((query.get(field) or "").lower() in (str(row[idx]).lower() if row[idx] else '') for idx, field in enumerate(query_fields)):
            results.append({
                'id': row[0],
                'firstName': row[1],
                'lastName': row[2],
                'badgeNumber': row[3],
                'title': row[4],
                'trainer': row[5],
                'date': row[6]
            })
    return results

def read_excel_sheet(sheet_name):
    try:
        return pd.read_excel(EXCEL_FILE_PATH, sheet_name=sheet_name)
    except Exception as e:
        logging.error(f"Failed to read {sheet_name} from Excel file: {e}")
        return None


# Debugging Token in Flask


def generate_token(user_id):
    try:
        token = jwt.encode({'id': user_id}, app.config['SECRET_KEY'])
        print(f"Generated Token: {token}")  # Debugging line
        return token
    except Exception as e:
        return None
# Debugging Home


@app.route('/api/users', methods=['GET'])
def api_users_options():  # Changed the function name here

    response = make_response()
    response.headers.add("Access-Control-Allow-Origin", "*")
    response.headers.add("Access-Control-Allow-Methods",
                         "GET,POST,PUT,DELETE,OPTIONS")
    response.headers.add("Access-Control-Allow-Headers",
                         "Content-Type, Authorization")
    return jsonify({"message": "This is a GET request"})


@app.route('/', methods=['GET'])
def home():
    return jsonify({"message": "Flask is running!"}), 200

# ADD ROUTES HERE

# Verify Token


@app.route('/verify-token', methods=['GET'])
@cross_origin()  # This will enable CORS for this route
@jwt_required()
def verify_token():
    return jsonify({"message": "Token is valid"}), 200


@app.route('/register', methods=['POST'])
@cross_origin()  # This will enable CORS for this route
def register():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    first_name = data.get('first_name')
    last_name = data.get('last_name')
    work_center = data.get('work_center')
    access_level = data.get('access_level', 1)  # Default to 1 if not provided

    hashed_password = pbkdf2_sha256.hash(password)

    users.insert_one({
        "username": username,
        "password": hashed_password,
        "first_name": first_name,
        "last_name": last_name,
        "work_center": work_center,
        "access_level": access_level,
    })

    return jsonify({"message": "User registered successfully"})


@app.route('/api/users', methods=['OPTIONS'])
def options_users():
    response = make_response()
    response.headers.add("Access-Control-Allow-Origin", "*")
    return response


@app.route('/login', methods=['POST'])
@cross_origin()  # This enables CORS for this route
def login():
    try:
        data = request.json
        username = data.get('username')
        password = data.get('password')

        # Fetch the user from MongoDB
        user = users.find_one({"username": username})

        # Check if user exists and the password is correct
        if user and pbkdf2_sha256.verify(password, user['password']):
            access_token = create_access_token(
                identity=username, expires_delta=timedelta(hours=5))
            return jsonify({"message": "Login successful", "access_token": access_token}), 200
        else:
            return jsonify({"message": "Invalid credentials"}), 401
    except Exception as e:
        print(f"An exception occurred: {e}")
        return jsonify({"message": "An error occurred while processing your request."}), 500


# Function to update Excel file


def update_excel_with_new_training(training):
    wb = load_workbook(os.getenv('EXCEL_FILE_PATH'))
    ws = wb['EmployeeSheet1']
    new_row = [
        training['firstName'],
        training['lastName'],
        training['badgeNumber'],
        training.get('title', ""),  # Training Title
        training.get('trainer', ""),  # Trainer
        training['date'],
        ""  # Empty cell for Work Center
    ]
    ws.append(new_row)
    wb.save(os.getenv('EXCEL_FILE_PATH'))


@app.route('/api/addNewTraining', methods=['POST'])
def add_new_training():
    try:
        data = request.json
        logging.debug(f"Received data: {data}")

        new_training = Training(
            # Using get to avoid KeyError
            firstName=data.get('firstName', None),
            lastName=data.get('lastName', None),
            badgeNumber=data.get('badgeNumber', None),
            title=data.get('title', None),  # New field
            date=data.get('date', None),  # New field
            trainer=data.get('trainer', None)  # New field
        )
        # Validation before adding to database
        if None in [new_training.firstName, new_training.lastName, new_training.badgeNumber, new_training.title,
                    new_training.date, new_training.trainer]:
            logging.error("One or more required fields are missing.")
            return jsonify({"message": "One or more required fields are missing."}), 422

        sql_db.session.add(new_training)
        sql_db.session.commit()

        return jsonify({"message": "New training added"}), 201

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        return jsonify({"message": "An error occurred while processing your request."}), 422


@app.route('/api/training-verification', methods=['GET'])
@cross_origin()  # This will enable CORS for this route
# @jwt_required()
def get_training_data():
    all_trainings = Training.query.all()
    return jsonify([training.serialize for training in all_trainings]), 200


@app.route('/search', methods=['GET'])
def search_training():
    query = request.args.get('query')
    df = pd.read_excel(EXCEL_FILE_PATH)
    filtered_df = df[df['Training Title'].str.contains(
        query, case=False, na=False)]
    results = filtered_df.to_dict('records')
    return jsonify(results)


@app.route('/getUsername', methods=['GET'])
@cross_origin()  # This will enable CORS for this route
@jwt_required()
def get_username():
    current_user = get_jwt_identity()
    user = users.find_one({"username": current_user})
    if user:
        return jsonify({"username": current_user, "first_name": user.get("first_name")}), 200
    else:
        return jsonify({"error": "User not found"}), 404


@app.route('/api/users/level1_and_2', methods=['GET'])
@cross_origin()  # This will enable CORS for this route
@jwt_required()
def get_level1_and_2_users():
    current_user = get_jwt_identity()
    current_user_data = users.find_one({"username": current_user})

    print(f"Current User: {current_user}")  # Debugging line
    print(f"Current User Data: {current_user_data}")  # Debugging line

    if current_user_data and current_user_data['access_level'] == 3:
        level1_and_2_users = users.find({"access_level": {"$in": [1, 2]}})
        level1_and_2_users_list = []
        for user in level1_and_2_users:
            user.pop('_id', None)
            level1_and_2_users_list.append(user)
        return jsonify({"success": True, "users": level1_and_2_users_list}), 200
    else:
        return jsonify({"success": False, "message": "Unauthorized"}), 401


@app.route('/api/users/update_access', methods=['POST'])
@cross_origin()  # This will enable CORS for this route
@jwt_required()
def update_user_access():
    data = request.json
    target_username = data.get("username")
    new_access_level = data.get("new_access_level")
    current_user = get_jwt_identity()
    current_user_data = users.find_one({"username": current_user})
    if current_user_data['work_center'] != 'Management' and new_access_level == 3:
        return jsonify({"success": False, "message": "Granting level 3 access requires a manager to approve. Please notify quality management to grant this access."}), 403

    if current_user_data['access_level'] == 3:
        result = users.update_one(
            {"username": target_username},
            {"$set": {"access_level": new_access_level}}
        )
        if result.modified_count > 0:
            return jsonify({"success": True, "message": "User access level updated successfully"}), 200
        else:
            return jsonify({"success": False, "message": "Failed to update user access level"}), 400
    else:
        return jsonify({"success": False, "message": "You do not have permission to change user access levels"}), 403


# CRUD API Routes

@app.route('/api/trainings', methods=['POST'])
@cross_origin()
@jwt_required()
def add_training():
    try:
        current_user = get_jwt_identity()
        data = request.json

        if None in [data.get(field) for field in ['firstName', 'lastName', 'badgeNumber', 'title', 'date', 'trainer']]:
            return jsonify({"message": "One or more required fields are missing."}), 422

        new_training = Training(**data)
        sql_db.session.add(new_training)
        sql_db.session.commit()

        return jsonify({"message": "New training added"}), 201

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        return jsonify({"message": "An error occurred while processing your request."}), 422


@app.route('/api/trainings/<int:id>', methods=['PUT'])
def update_training(id):
    try:
        training = Training.query.get(id)
        if not training:
            return jsonify({"message": "Training not found"}), 404
        data = request.json
        for key, value in data.items():
            setattr(training, key, value)
        sql_db.session.commit()
        return jsonify({"message": "Training updated"}), 200
    except Exception as e:
        return jsonify({"message": f"An error occurred: {str(e)}"}), 500


@app.route('/api/trainings/<int:id>', methods=['DELETE'])
def delete_training(id):
    try:
        training = Training.query.get(id)
        if not training:
            return jsonify({"message": "Training not found"}), 404
        sql_db.session.delete(training)
        sql_db.session.commit()
        return jsonify({"message": "Training deleted"}), 200
    except Exception as e:
        return jsonify({"message": f"An error occurred: {str(e)}"}), 500

# GET USER NAME FOR SMART SHEET


@app.route('/api/getUserDetails', methods=['GET'])
@cross_origin()
@jwt_required()
def get_user_details():
    try:
        print("Inside getUserDetails API")
        current_user = get_jwt_identity()  # This should return the username
        print(f"Current User: {current_user}")

        if current_user:
            # Updated to use 'username'
            user = users.find_one({"username": current_user})
            if user:
                print(f"Found user: {user}")
                return jsonify({"name": f"{user['first_name']} {user['last_name']}"})
            else:
                print("User not found in database")
        else:
            print("No current user")

        return jsonify({"error": "User not found"}), 404
    except Exception as e:
        print(f"An exception occurred: {e}")
        return jsonify({"error": "An error occurred"}), 500


@app.route('/api/export_to_excel', methods=['POST'])
def export_to_excel():
    try:
        print("Received Data:", request.json)
        data = request.json

        excel_path = os.getenv('EXCEL_FILE_PATH')
        wb = load_workbook(excel_path)
        ws = wb.active

        for training in data:
            ws.append([training.get(field, "") for field in [
                      "firstName", "lastName", "badgeNumber", "title", "trainer", "date", ""]])

        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(excel_path)
        return jsonify({"message": "Training data exported to Excel"}), 200

    except Exception as e:
        traceback.print_exc()
        logging.error(f"An error occurred: {e}")
        return jsonify({"message": "An error occurred while processing your request.", "error": str(e)}), 422


if __name__ == '__main__':
    with app.app_context():
        sql_db.create_all()  # Create all tables for SQLAlchemy
    app.run(debug=True, port=5000)


# Route for checking training stats by department
@app.route('/api/check_training', methods=['GET'])
def check_training():
    try:
        employee_sheet = read_excel_sheet('EmployeeSheet1')
        required_procedures = read_excel_sheet('Required Procedures')

        if employee_sheet is None or required_procedures is None:
            return jsonify({'data': None, 'error': 'Failed to read Excel sheets'}), 500

        current_date = datetime.now()
        one_year_ago = current_date - timedelta(days=365)

        department_training_stats = defaultdict(
            lambda: {'completed': 0, 'due': 0})

        for index, row in required_procedures.iterrows():
            work_instruction = row['Work Instructions']
            department = row['Department']

            relevant_records = employee_sheet[
                (employee_sheet['Training Title'] == work_instruction) &
                (employee_sheet['Department'] == department)
            ]

            completed_within_year = any(pd.to_datetime(
                relevant_records['Date']) > one_year_ago)

            if completed_within_year:
                department_training_stats[department]['completed'] += 1
            else:
                department_training_stats[department]['due'] += 1

        return jsonify({'data': department_training_stats, 'error': None}), 200
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        return jsonify({'data': None, 'error': str(e)}), 500


@app.route('/api/total_trained_employees_v2', methods=['GET'])
def total_trained_employees_v2():
    try:
        employee_sheet = read_excel_sheet('EmployeeSheet1')
        if employee_sheet is None:
            return jsonify({'error': 'Failed to read EmployeeSheet1'}), 500

        employee_data = read_excel_sheet('EmployeeData')
        if employee_data is None:
            return jsonify({'error': 'Failed to read EmployeeData'}), 500

        total_trained = defaultdict(int)

        for department in employee_data['Department'].unique():
            trained_employees = employee_sheet[employee_sheet['Department'] == department]
            total_trained[department] = len(
                trained_employees['Badge Number'].unique())

        return jsonify(total_trained), 200

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/check_employee_training', methods=['GET'])
def check_employee_training():
    try:
        employee_sheet = pd.read_excel(
            EXCEL_FILE_PATH, sheet_name='EmployeeSheet1')
        required_procedures = pd.read_excel(
            EXCEL_FILE_PATH, sheet_name='Required Procedures')

        current_date = datetime.now()
        one_year_ago = current_date - timedelta(days=365)

        department_training_stats = defaultdict(
            lambda: {'completed': 0, 'due': 0})

        for index, row in required_procedures.iterrows():
            work_instruction = row['Work Instructions']
            department = row['Department']

            relevant_records = employee_sheet[
                (employee_sheet['Training Title'] == work_instruction) &
                (employee_sheet['Department'] == department)
            ]

            completed_within_year = any(
                pd.to_datetime(relevant_records['Date']) > one_year_ago
            )

            if completed_within_year:
                department_training_stats[department]['completed'] += 1
            else:
                department_training_stats[department]['due'] += 1

        return jsonify(department_training_stats), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/total_trained_employees_new', methods=['GET'])
def total_trained_employees_new():
    try:
        excel_path = EXCEL_FILE_PATH
        employee_sheet = pd.read_excel(excel_path, sheet_name='EmployeeSheet1')
        employee_data = pd.read_excel(excel_path, sheet_name='EmployeeData')

        total_trained = defaultdict(int)

        for department in employee_data['Department'].unique():
            trained_employees = employee_sheet[employee_sheet['Department'] == department]
            total_trained[department] = len(
                trained_employees['Badge Number'].unique())

        return jsonify(total_trained), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    
@app.route('/api/get_combined_employee_data', methods=['GET'])
def get_combined_employee_data():
    try:
        query = {
            "firstName": request.args.get('firstName', ''),
            "lastName": request.args.get('lastName', ''),
            "badgeNumber": request.args.get('badgeNumber', '')
        }
        data = read_from_excel(query)
        return jsonify(data), 200
    except Exception as e:
        logging.error(f"Error: {e}")
        return jsonify({"error": "Internal Server Error"}), 500